import openpyxl as xl  # Import the openpyxl library as xl for working with Excel files

wb = xl.load_workbook('Timeseddel.xlsx')  # Load the workbook named 'Timeseddel.xlsx'

sheet = wb[wb.sheetnames[0]] #selects the first Excel sheet 
print(sheet.title) # Print the name of the specific sheet in the workbook
sheet.title = 'Title lavet af Python'
print(sheet.title)

print(sheet.max_row)
print(sheet.max_column)

wb.create_sheet(title='ark lavet af Python', index=1)
print(wb.sheetnames)

print(sheet['B8'].value)
sheet['B8'] = 'I beg your pardon, I never promised you a rose garden'
print(sheet['B8'].value)
print(sheet.cell(column=1, row=1).value)
sheet.sheet_properties.tabColor = "1072BA"
wb.save('TimeseddelJuli.xlsx')