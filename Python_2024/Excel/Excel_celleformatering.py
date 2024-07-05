import openpyxl as xl  # Import the openpyxl library as xl for working with Excel files
import datetime as dt  # Import the datetime module as dt for date and time operations
import numbers  # Import the numbers module for type checking
import sys
from openpyxl.styles import PatternFill, Border, Side, Font

try:
    wb = xl.load_workbook(sys.argv[1])
except FileNotFoundError as e:
    print(e)
except NameError as e:
    print('Fejl')
    exit()

 
print() #should print the name of the Excel file
# Load the workbook named 'Timeseddel.xlsx'
wb = xl.load_workbook('Timesheet.xlsx')

# Access the first sheet in the workbook
sheet = wb[wb.sheetnames[0]]

# Get the current date and time
today = dt.datetime.today()

fromDate = 2  # Start checking from row 2
# Loop to find the row where the date matches today's date
while today.date() != dt.datetime(sheet.cell(column=1, row=fromDate).value):
    fromDate += 1
    
toDate = fromDate  # Set toDate as fromDate

# Variable to accumulate sum of column 6 values
sum = 0

# Loop to sum the values in column 6 as long as column 8 contains a non-string value
while not(isinstance(sheet.cell(column=8, row=fromDate).value, str)):
    fromDate -= 1
    if isinstance(sheet.cell(column=6, row=fromDate).value, numbers.Real):
        sum += sheet.cell(column=6, row=fromDate).value

sheet['E1'] = 'Udbetalt'
sheet['E1'].font = Font(name='Calibri', size=11, bold=True, italic=False)
sheet.cell(column=8, row=toDate).value = sum 
sheet.cell(column=8, row=toDate).fill = PatternFill(fill_type='solid', start_color='00FF00')
sheet.cell(column=8, row=toDate).border = Border(bottom=Side(border_style='thick'))
# Save the workbook with a new name 'Timesheet.xlsx'
wb.save('Timesheet.xlsx')
